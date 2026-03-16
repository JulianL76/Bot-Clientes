import pyautogui
import pandas as pd
import time
import sys
import os
import threading
import pygetwindow as gw
import winsound
import unicodedata
import openpyxl

# ==========================================
# CONFIGURACIÓN INICIAL
# ==========================================
EXCEL_FILE = 'datos.xlsx'
SHEET_NAME = 0 

# Pausa estándar básica
pyautogui.PAUSE = 0.2
pyautogui.FAILSAFE = True

# ==========================================
# VARIABLES GLOBALES
# ==========================================
class BotState:
    pause_requested = False
    digito_requested = False
    creado_requested = False
    current_row_index = 0
    main_window_title = ""
    is_saving = False

# ==========================================
# FUNCIONES AUXILIARES
# ==========================================

def limpiar_texto(texto):
    """
    Elimina saltos de línea, tabulaciones y cualquier caracter de control
    de un string para que pyautogui no los interprete como teclas especiales.
    """
    if not texto:
        return ""
    texto = str(texto)
    # Reemplazar saltos de línea y tabs por espacio
    texto = texto.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
    # Eliminar caracteres de control (categoría Unicode 'C' excepto espacio normal)
    texto = ''.join(c for c in texto if not unicodedata.category(c).startswith('C'))
    # Colapsar múltiples espacios en uno solo y quitar espacios al inicio/final
    texto = ' '.join(texto.split())
    return texto

def monitor_de_popups():
    """
    Hilo en segundo plano que vigila constantemente la ventana activa.
    """
    while True:
        try:
            active_window = gw.getActiveWindow()
            if active_window and BotState.main_window_title and not BotState.is_saving:
                current_title = active_window.title
                
                if current_title and current_title != BotState.main_window_title and "abako" not in current_title.lower() and "cmd.exe" not in current_title.lower() and "python" not in current_title.lower():
                    if "digito" in current_title.lower() or "dígito" in current_title.lower():
                        if not BotState.digito_requested:
                            print(f"\n    [!] Detectado popup: '{current_title}'. Solicitando escape automático...")
                            BotState.digito_requested = True
                    elif "planilla nuevo cliente" in current_title.lower() or "recuperar su" in current_title.lower():
                        if not BotState.creado_requested:
                            print(f"\n    [!] Detectado popup de cliente existente: '{current_title}'. Solicitando escape automático...")
                            BotState.creado_requested = True
                    else:
                        if not BotState.pause_requested:
                            print("\n" + "!"*60)
                            print(f"[ALERTA] Se detectó una ventana emergente inesperada: '{current_title}'")
                            print(f"[ALERTA] Fila actual: {BotState.current_row_index}.")
                            print("[ALERTA] Escribe 's' y presiona ENTER para continuar.")
                            print("!"*60 + "\n")
                            BotState.pause_requested = True
                            for _ in range(3):
                                winsound.Beep(1000, 400)
                                time.sleep(0.1)
        except:
            pass
        time.sleep(0.1)

def check_pause():
    if BotState.digito_requested: raise Exception("DIGITO_VERIFICACION")
    if BotState.creado_requested: raise Exception("CLIENTE_YA_CREADO")
    if BotState.pause_requested: raise Exception("PAUSA_SEGURIDAD")

def smart_sleep(duration):
    slices = int(duration / 0.1)
    for _ in range(slices):
        check_pause()
        time.sleep(0.1)

# ==========================================
# LÓGICA DEL BOT
# ==========================================

def read_data_from_excel(file_path):
    try:
        if not os.path.exists(file_path):
            print(f"[-] Error: El archivo {file_path} no existe.")
            return None
        print(f"[*] Leyendo archivo Excel: {file_path}...")
        df = pd.read_excel(file_path, sheet_name=SHEET_NAME)
        print(f"[+] Se cargaron {len(df)} filas correctamente.")
        return df
    except Exception as e:
        print(f"[-] Error al leer el archivo Excel: {e}")
        return None

def informacion_basica(row):
    """
    Llena la sección de Información Básica REVERTIDA al bot original.
    """
    print("\n--- [ INFORMACIÓN BÁSICA ] ---")
    
    # EXTRACCIÓN Y LIMPIEZA DE DATOS DE EXCEL
    dato_razon_social   = limpiar_texto(str(row.get('Razon Social', '')))
    dato_nombre_comun   = limpiar_texto(str(row.get('Nombre Común', '')))
    dato_identificacion = limpiar_texto(str(row.get('Identificación', '')))
    if dato_identificacion.endswith('.0'): dato_identificacion = dato_identificacion[:-2]
    dato_canal_texto    = limpiar_texto(str(row.get('Canal', '')))
    dato_segmento_texto = limpiar_texto(str(row.get('Segmento', '')))
    dato_asesor_texto   = limpiar_texto(str(row.get('Asesor', '')))
    dato_zona_texto     = limpiar_texto(str(row.get('Zona', '')))

    if dato_razon_social.lower()   == 'nan': dato_razon_social   = ""
    if dato_nombre_comun.lower()   == 'nan': dato_nombre_comun   = ""
    if dato_identificacion.lower() == 'nan': dato_identificacion = ""
    if dato_canal_texto.lower()    == 'nan': dato_canal_texto    = ""
    if dato_segmento_texto.lower() == 'nan': dato_segmento_texto = ""
    if dato_asesor_texto.lower()   == 'nan': dato_asesor_texto   = ""
    if dato_zona_texto.lower()     == 'nan': dato_zona_texto     = ""

    # Lógica Canal
    dato_canal = ""
    canal_upper = dato_canal_texto.upper()
    if "SUPERMERCADO" in canal_upper: dato_canal = "sss"
    elif "MINORISTA" in canal_upper: dato_canal = "mm"
    elif "MAYORISTA" in canal_upper: dato_canal = "m"
    elif "TIENDA" in canal_upper: dato_canal = "t"
    elif canal_upper: dato_canal = canal_upper[0]
    
    # Lógica Segmento
    dato_segmento = ""
    segmento_upper = dato_segmento_texto.upper()
    if "SUPERMERCADO" in segmento_upper: dato_segmento = "ssss"
    elif "MINORISTA" in segmento_upper: dato_segmento = "mm"
    elif "MAYORISTA" in segmento_upper: dato_segmento = "m"
    elif "TIENDA" in segmento_upper: dato_segmento = "t"
    elif segmento_upper: dato_segmento = segmento_upper[0]

    # Paso 1: Clic inicial en Razón Social y escribir
    print("    - Paso 1: Escribiendo Razón Social")
    time.sleep(0.5)
    pyautogui.moveTo(515, 115, duration=0.5)
    pyautogui.click(x=515, y=115)
    pyautogui.click(x=515, y=115)
    time.sleep(0.5)
    pyautogui.write(dato_razon_social)
    time.sleep(1.0)

    # Paso 2: ENTER+ENTER → Nombre Común y escribir
    print("    - Paso 2: Presionando ENTER y escribiendo Nombre Común")
    pyautogui.press("enter")
    time.sleep(0.2)
    pyautogui.press("enter")
    time.sleep(0.8)
    pyautogui.write(dato_nombre_comun)
    time.sleep(2.0)

    # Paso 3: ENTER+ENTER → Tipo Documento (Cédula)
    print("    - Paso 3: Seleccionando Cédula (ENTER+ENTER + C + Enter)")
    pyautogui.press("enter")
    time.sleep(0.2)
    pyautogui.press("enter")
    time.sleep(1.0)
    pyautogui.press("c")
    time.sleep(1.5)
    pyautogui.press("enter")
    time.sleep(1.0)

    # Paso 4: Escribir Identificación
    print("    - Paso 4: Escribiendo Identificación")
    time.sleep(1.0)
    pyautogui.write(dato_identificacion)
    time.sleep(0.5)
    pyautogui.press("tab")
    
    print("    - Paso 4.1: Esperando popups post-identificación...")
    smart_sleep(1.0)
    check_pause()
    
    # Paso 5: Asesor
    print(f"    - Paso 5: Asesor ({dato_asesor_texto})")
    pyautogui.press("tab")
    time.sleep(0.5)
    if dato_asesor_texto: 
        asesor_upper = dato_asesor_texto.upper()
        if "MARISOL VARGAS GOMEZ" in asesor_upper:
            pyautogui.write("mar")
            time.sleep(0.5)
            for _ in range(4):
                pyautogui.press("down")
                time.sleep(0.1)
        elif "GABRIEL MAURICIO GARCIA URREA" in asesor_upper:
            pyautogui.write("ga")
        elif "CARLOS ANDRES MEJIA CARREÑO" in asesor_upper:
            pyautogui.write("ca")
        elif "ZAIDE LEONOR VASQUEZ ARIAS" in asesor_upper:
            pyautogui.write("za")
        elif "SANDRA PAOLA CORZO JURADO" in asesor_upper:
            pyautogui.write("san")
            time.sleep(0.5)
            pyautogui.press("down")
            time.sleep(0.1)
        else:
            pyautogui.write(dato_asesor_texto[:3])
    pyautogui.press("enter")
    time.sleep(0.3)

    # Paso 6: Canal
    print(f"    - Paso 6: Canal ({dato_canal_texto})")
    time.sleep(0.5)
    if dato_canal: pyautogui.write(dato_canal, interval=0.15)
    pyautogui.press("enter")
    time.sleep(0.3)
    
    # Paso 7: Segmento
    print(f"    - Paso 7: Segmento ({dato_segmento_texto})")
    time.sleep(0.5)
    if dato_segmento: pyautogui.write(dato_segmento, interval=0.15)
    pyautogui.press("enter")
    time.sleep(0.3)

    # Paso 8: Zona
    print(f"    - Paso 8: Zona ({dato_zona_texto})")
    time.sleep(0.5)
    if dato_zona_texto: pyautogui.write(dato_zona_texto, interval=0.15)
    pyautogui.press("enter")
    time.sleep(0.3)

def facturacion(row):
    """
    Llena la sección de Facturación.
    """
    print("\n--- [ FACTURACIÓN ] ---")
    
    dato_forma_pago = limpiar_texto(str(row.get('Forma de Pago', '')))
    dato_cupo = limpiar_texto(str(row.get('Cupo', '')))
    dato_cuotas = limpiar_texto(str(row.get('Cuotas', '')))
    dato_vencimiento = limpiar_texto(str(row.get('Vencimiento', '')))
    dato_lista_precios = limpiar_texto(str(row.get('Lista de Precios', '')))
    
    if dato_forma_pago.lower() == 'nan': dato_forma_pago = ""
    if dato_cupo.lower() == 'nan': dato_cupo = ""
    if dato_cuotas.lower() == 'nan': dato_cuotas = ""
    if dato_vencimiento.lower() == 'nan': dato_vencimiento = ""
    if dato_lista_precios.lower() == 'nan': dato_lista_precios = ""
    
    dato_lista_corto = dato_lista_precios[:3] if dato_lista_precios else ""

    print("    - Paso 1: Clic Inicial (389, 414)")
    pyautogui.moveTo(389, 414, duration=0.5)
    pyautogui.click(x=389, y=414)
    smart_sleep(0.1)
    
    if "CONTADO" not in dato_forma_pago.upper():
        print("    - Paso 2: Clic en Forma de Pago Crédito (674, 458)")
        pyautogui.moveTo(674, 458, duration=0.5)
        pyautogui.click(x=674, y=458)
        smart_sleep(0.1)
        
        pyautogui.write("c")
        smart_sleep(0.1)
        pyautogui.press("enter")
        smart_sleep(0.1)
        
        print(f"    - Paso 3: Escribiendo Cupo ({dato_cupo}) y Enter x2")
        pyautogui.write(dato_cupo)
        smart_sleep(0.1)
        pyautogui.press("enter")
        smart_sleep(0.1)
        pyautogui.press("enter")
        smart_sleep(0.1)
        
        print(f"    - Paso 4: Escribiendo Cuotas ({dato_cuotas}) y Enter x2")
        pyautogui.write(dato_cuotas)
        smart_sleep(0.1)
        pyautogui.press("enter")
        smart_sleep(0.1)
        
        # Lógica de vencimiento según días
        vencimiento_num = "".join(c for c in dato_vencimiento if c.isdigit())
        mapeo_vencimiento = {
            "7": ("semana", "semsem"),
            "14": ("quincenal", "qu"),
            "21": ("20 días", "20"),
            "28": ("25 días", "25"),
            "30": ("mensual", "me"),
            "45": ("cuarenta", "cu")
        }
        
        if vencimiento_num in mapeo_vencimiento:
            label, teclas = mapeo_vencimiento[vencimiento_num]
            print(f"    - Paso 4.1: Vencimiento {vencimiento_num} días ({label}) detectado. Escribiendo '{teclas}' + Enter.")
            pyautogui.write(teclas)
            smart_sleep(0.1)
            pyautogui.press("enter")
            smart_sleep(0.1)
        elif vencimiento_num:
            # Caso por defecto si hay un número pero no está en el mapeo
            print(f"    - Paso 4.1: Vencimiento {vencimiento_num} días detectado (sin mapeo específico).")
            pyautogui.write(vencimiento_num)
            smart_sleep(0.1)
            pyautogui.press("enter")
            smart_sleep(0.1)
    else:
        print("    --> Forma de pago Contado detectada. Omitiendo pasos de crédito.")

    check_pause()
    print(f"    - Paso 5: Clic en Lista de Precios ({dato_lista_precios}) -> ({dato_lista_corto} + Enter)")
    pyautogui.moveTo(1054, 461, duration=0.5)
    pyautogui.click(x=1054, y=461)
    smart_sleep(0.2)
    
    if dato_lista_corto:
        pyautogui.write(dato_lista_corto, interval=0.15)
    smart_sleep(0.1)
    pyautogui.press("enter")
    smart_sleep(0.2)

def contacto(row):
    """
    Llena la sección de Contacto.
    """
    print("\n--- [ CONTACTO ] ---")
    
    dato_nombre_contacto = limpiar_texto(str(row.get('Nombre Contacto', '')))
    dato_apellido_contacto = limpiar_texto(str(row.get('Apellido Contacto', '')))
    dato_identificacion_contacto = limpiar_texto(str(row.get('Identificación Contacto', '')))
    if str(dato_identificacion_contacto).endswith('.0'): dato_identificacion_contacto = dato_identificacion_contacto[:-2]
    
    if dato_nombre_contacto.lower() == 'nan': dato_nombre_contacto = ""
    if dato_apellido_contacto.lower() == 'nan': dato_apellido_contacto = ""
    if dato_identificacion_contacto.lower() == 'nan': dato_identificacion_contacto = ""

    # Paso 1: Clic Normal en Contacto
    check_pause()
    print("    - Paso 1: Clic Inicial (491, 416)")
    pyautogui.moveTo(491, 416, duration=0.5)
    pyautogui.click(x=491, y=416)
    smart_sleep(0.1)
    
    # Escribe nombre contacto y Enter
    print(f"    - Paso 2: Borrando y Escribiendo Nombre Contacto ({dato_nombre_contacto}) y Presionando Enter")
    for _ in range(4):
        pyautogui.hotkey('ctrl', 'delete')
    pyautogui.write(dato_nombre_contacto)
    smart_sleep(0.1)
    pyautogui.press("enter")
    smart_sleep(0.1)
    
    # Escribe apellidos contacto y enter
    print(f"    - Paso 3: Borrando y Escribiendo Apellido Contacto ({dato_apellido_contacto}) y Presionando Enter")
    for _ in range(4):
        pyautogui.hotkey('ctrl', 'delete')
    pyautogui.write(dato_apellido_contacto)
    smart_sleep(0.1)
    pyautogui.press("enter")
    smart_sleep(0.1)
    
    # Escribe indentificacion contacto
    print(f"    - Paso 4: Escribiendo Identificación Contacto ({dato_identificacion_contacto})")
    pyautogui.write(dato_identificacion_contacto)
    smart_sleep(0.1)

def direccion(row):
    """
    Llena la sección de Dirección.
    """
    print("\n--- [ DIRECCIÓN ] ---")
    
    dato_departamento = limpiar_texto(str(row.get('Departamento', '')))
    dato_ciudad = limpiar_texto(str(row.get('Ciudad', '')))
    dato_barrio = limpiar_texto(str(row.get('Barrio', '')))
    dato_prefijo = limpiar_texto(str(row.get('PREFIJO', '')))
    dato_parte_uno = limpiar_texto(str(row.get('PARTE UNO', '')))
    dato_parte_dos = limpiar_texto(str(row.get('PARTE DOS', '')))
    dato_parte_tres = limpiar_texto(str(row.get('PARTE TRES', '')))
    
    if dato_departamento.lower() == 'nan': dato_departamento = ""
    if dato_ciudad.lower() == 'nan': dato_ciudad = ""
    if dato_barrio.lower() == 'nan': dato_barrio = ""
    if dato_prefijo.lower() == 'nan': dato_prefijo = ""
    if dato_parte_uno.lower() == 'nan': dato_parte_uno = ""
    if dato_parte_dos.lower() == 'nan': dato_parte_dos = ""
    if dato_parte_tres.lower() == 'nan': dato_parte_tres = ""

    # Lógica de Departamento (2 letras)
    dato_departamento_corto = dato_departamento[:2] if dato_departamento else ""
    
    # Lógica de Ciudad (3 letras)
    dato_ciudad_corto = dato_ciudad[:3] if dato_ciudad else ""

    # Lógica de Barrio (3 letras)
    dato_barrio_corto = dato_barrio[:3] if dato_barrio else ""
    
    print("    - Paso 1: Clic (556, 408)")
    check_pause()
    pyautogui.moveTo(556, 408, duration=0.5)
    pyautogui.click(x=556, y=408)
    smart_sleep(0.2)
    
    print("    - Paso 2: Clic en Departamento (398, 501)")
    pyautogui.moveTo(398, 501, duration=0.5)
    pyautogui.click(x=398, y=501)
    smart_sleep(0.2)
    
    print(f"    - Paso 2.1: Escribiendo Departamento ({dato_departamento_corto}) + Enter")
    if dato_departamento_corto:
        pyautogui.write(dato_departamento_corto, interval=0.15)
    smart_sleep(0.2)
    pyautogui.press("enter")
    smart_sleep(0.2)
    
    print(f"    - Paso 3: Escribiendo Ciudad ({dato_ciudad_corto}) + Enter")
    if dato_ciudad_corto:
        pyautogui.write(dato_ciudad_corto, interval=0.15)
    smart_sleep(0.2)
    pyautogui.press("enter")
    smart_sleep(0.2)

    if dato_barrio.upper() == "CENTRO":
        print(f"    - Paso 3.1: Escribiendo Barrio COMPLETO rápido ({dato_barrio}) + Enter")
        pyautogui.write(dato_barrio, interval=0.01)
    elif dato_barrio.upper() == "CIUDAD VALENCIA":
        print(f"    - Paso 3.1: Caso especial CIUDAD VALENCIA. Escribiendo 'CIU' + 3 Down")
        pyautogui.write("CIU", interval=0.15)
        smart_sleep(0.2)
        for _ in range(3):
            pyautogui.press("down")
            time.sleep(0.1)
    elif dato_barrio.upper() == "CIUDADELA REAL DE MINAS":
        print(f"    - Paso 3.1: Caso especial CIUDADELA REAL DE MINAS. Escribiendo 'CIU' + 9 Down")
        pyautogui.write("CIU", interval=0.15)
        smart_sleep(0.2)
        for _ in range(9):
            pyautogui.press("down")
            time.sleep(0.1)
    elif dato_barrio.upper() == "CONUCOS":
        print(f"    - Paso 3.1: Caso especial CONUCOS. Escribiendo 'CON' + 6 Down")
        pyautogui.write("CON", interval=0.15)
        smart_sleep(0.2)
        for _ in range(6):
            pyautogui.press("down")
            time.sleep(0.1)
    elif "DIAMANTE II" in dato_barrio.upper() or "DIAMANTE 2" in dato_barrio.upper():
        print(f"    - Paso 3.1: Caso especial DIAMANTE II. Escribiendo 'DIA' + 2 Down")
        pyautogui.write("DIA", interval=0.15)
        smart_sleep(0.2)
        for _ in range(2):
            pyautogui.press("down")
            time.sleep(0.1)
    elif dato_barrio.upper() == "EL CRISTAL":
        print(f"    - Paso 3.1: Caso especial EL CRISTAL. Escribiendo 'EL' + 5 Down")
        pyautogui.write("EL", interval=0.15)
        smart_sleep(0.2)
        for _ in range(5):
            pyautogui.press("down")
            time.sleep(0.1)
    elif dato_barrio.upper() == "EL DIVISO":
        print(f"    - Paso 3.1: Caso especial EL DIVISO. Escribiendo 'EL' + 4 Down")
        pyautogui.write("EL", interval=0.15)
        smart_sleep(0.2)
        for _ in range(4):
            pyautogui.press("down")
            time.sleep(0.1)
    elif dato_barrio.upper() == "EL DORADO":
        print(f"    - Paso 3.1: Caso especial EL DORADO. Escribiendo 'EL' + 2 Down")
        pyautogui.write("EL", interval=0.15)
        smart_sleep(0.2)
        for _ in range(2):
            pyautogui.press("down")
            time.sleep(0.1)
    elif dato_barrio.upper() == "SAN FRANCISCO":
        print(f"    - Paso 3.1: Caso especial SAN FRANCISCO. Escribiendo 'SAN' + 4 Down")
        pyautogui.write("SAN", interval=0.15)
        smart_sleep(0.2)
        for _ in range(4):
            pyautogui.press("down")
            time.sleep(0.1)
    elif dato_barrio.upper() == "SAN GERARDO":
        print(f"    - Paso 3.1: Caso especial SAN GERARDO. Escribiendo 'SAN' + 7 Down")
        pyautogui.write("SAN", interval=0.15)
        smart_sleep(0.2)
        for _ in range(7):
            pyautogui.press("down")
            time.sleep(0.1)
    elif dato_barrio.upper() == "SAN LUIS":
        print(f"    - Paso 3.1: Caso especial SAN LUIS. Escribiendo 'SAN' + 11 Down")
        pyautogui.write("SAN", interval=0.15)
        smart_sleep(0.2)
        for _ in range(11):
            pyautogui.press("down")
            time.sleep(0.1)
    elif dato_barrio.upper() == "SAN MARTIN" or dato_barrio.upper() == "SAN MARTÍN":
        print(f"    - Paso 3.1: Caso especial SAN MARTIN. Escribiendo 'SAN' + 14 Down")
        pyautogui.write("SAN", interval=0.15)
        smart_sleep(0.2)
        for _ in range(14):
            pyautogui.press("down")
            time.sleep(0.1)
    elif dato_barrio.upper() == "SAN MIGUEL":
        print(f"    - Paso 3.1: Caso especial SAN MIGUEL. Escribiendo 'SAN' + 19 Down")
        pyautogui.write("SAN", interval=0.15)
        smart_sleep(0.2)
        for _ in range(19):
            pyautogui.press("down")
            time.sleep(0.1)
    elif dato_barrio.upper() == "SAN PEDRO":
        print(f"    - Paso 3.1: Caso especial SAN PEDRO. Escribiendo 'SAN' + 21 Down")
        pyautogui.write("SAN", interval=0.15)
        smart_sleep(0.2)
        for _ in range(21):
            pyautogui.press("down")
            time.sleep(0.1)
    elif dato_barrio.upper() == "RICUARTE":
        print(f"    - Paso 3.1: Caso especial RICUARTE. Escribiendo solo 2 letras ('RI')")
        pyautogui.write("RI", interval=0.15)
    else:
        print(f"    - Paso 3.1: Escribiendo Barrio ({dato_barrio_corto}) + Enter")
        if dato_barrio_corto:
            pyautogui.write(dato_barrio_corto, interval=0.15)
    smart_sleep(0.2)
    pyautogui.press("enter")
    smart_sleep(0.2)
    
    print("    - Paso 4: Clic en tipo de vía (745, 477)")
    pyautogui.moveTo(745, 477, duration=0.5)
    pyautogui.click(x=745, y=477)
    smart_sleep(0.2)
    
    print(f"    - Paso 4.1: Escribiendo Prefijo ({dato_prefijo}) + Enter")
    if dato_prefijo:
        pyautogui.write(dato_prefijo, interval=0.15)
    smart_sleep(0.2)
    pyautogui.press("enter")
    smart_sleep(0.2)

    print(f"    - Paso 5: Escribiendo Dirección ({dato_parte_uno} - {dato_parte_dos} - {dato_parte_tres})")
    pyautogui.write(dato_parte_uno)
    pyautogui.press("tab")
    pyautogui.write(dato_parte_dos)
    pyautogui.press("tab")
    pyautogui.write(dato_parte_tres)
    smart_sleep(0.5)
    pyautogui.press("enter")
    smart_sleep(0.2)

def telefono(row):
    """
    Llena la sección de Teléfono.
    """
    print("\n--- [ TELÉFONO ] ---")
    
    dato_telefono = limpiar_texto(str(row.get('Teléfono', row.get('Telefono', ''))))
    if dato_telefono.lower() == 'nan': dato_telefono = ""
    if dato_telefono.endswith('.0'): dato_telefono = dato_telefono[:-2]

    # Paso 1: Clic en pestaña/sección
    check_pause()
    print("    - Paso 1: Clic en pestaña Teléfono (668, 417)")
    pyautogui.moveTo(668, 417, duration=0.5)
    pyautogui.click(x=668, y=417)
    smart_sleep(0.2)
    
    # Paso 2: Clic en botón (Nuevo/Añadir)
    print("    - Paso 2: Clic en botón (44, 55)")
    pyautogui.moveTo(44, 55, duration=0.5)
    pyautogui.click(x=44, y=55)
    smart_sleep(0.2)
    
    # Paso 3: Navegación con Tabs
    print("    - Paso 3: Presionando Tab x3")
    for _ in range(3):
        pyautogui.press("tab")
        smart_sleep(0.1)
    
    # Paso 4: READ_WRITE en 1171, 491
    print(f"    - Paso 4: Escribiendo Teléfono ({dato_telefono}) en (1171, 491)")
    pyautogui.moveTo(1171, 491, duration=0.5)
    pyautogui.click(x=1171, y=491)
    smart_sleep(0.2)
    pyautogui.write(dato_telefono)
    smart_sleep(0.2)
    pyautogui.press("enter")

def correo(row):
    """
    Llena la sección de Correo Electrónico.
    """
    print("\n--- [ CORREO ELECTRÓNICO ] ---")
    
    dato_correo = limpiar_texto(str(row.get('Correo', row.get('Email', ''))))
    if dato_correo.lower() == 'nan' or not dato_correo:
        print("    --> No hay correo electrónico para esta fila. Saltando.")
        return

    # Paso 1: Clic en pestaña Correo
    check_pause()
    print("    - Paso 1: Clic en pestaña Correo (756, 415)")
    pyautogui.moveTo(756, 415, duration=0.5)
    pyautogui.click(x=756, y=415)
    smart_sleep(0.2)
    
    # Paso 2: Clic en botón Nuevo/Añadir
    print("    - Paso 2: Clic en botón (45, 56)")
    pyautogui.moveTo(45, 56, duration=0.5)
    pyautogui.click(x=45, y=56)
    smart_sleep(0.2)
    
    # Paso 3: Presionar C, C (para seleccionar tipo)
    print("    - Paso 3: Presionando 'c' x2 + Enter")
    pyautogui.press("c")
    smart_sleep(0.1)
    pyautogui.press("c")
    smart_sleep(0.1)
    pyautogui.press("enter")
    smart_sleep(0.2)
    
    # Paso 4: Tab para ir al campo de texto
    print("    - Paso 4: Presionando Tab")
    pyautogui.press("tab")
    smart_sleep(0.1)
    
    # Paso 5: Escribir el correo
    pyautogui.write(dato_correo)
    smart_sleep(0.2)
    pyautogui.press("enter")

def marcar_completado(excel_row_num, valor='X'):
    """
    Marca la fila en el Excel como procesada.
    """
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active if isinstance(SHEET_NAME, int) else wb[SHEET_NAME]
        
        col_completado = None
        for cell in ws[1]:
            if cell.value and str(cell.value).strip().lower() == 'completado':
                col_completado = cell.column
                break
        
        if col_completado is None:
            col_completado = ws.max_column + 1
            ws.cell(row=1, column=col_completado, value='Completado')
        
        ws.cell(row=excel_row_num, column=col_completado, value=valor)
        wb.save(EXCEL_FILE)
        print(f"    [Excel] ✅ Fila {excel_row_num} marcada en el Excel.")
    except Exception as e:
        print(f"    [Excel] ⚠️ Error al marcar Excel: {e}")

def guardar_y_continuar():
    """
    Realiza el proceso de guardado y limpia para el siguiente.
    """
    print("\n--- [ GUARDAR Y CONTINUAR ] ---")
    BotState.is_saving = True
    
    # 1. Clic Guardar
    print("    - Paso 1: Clic Guardar (93, 52)")
    pyautogui.moveTo(93, 52, duration=0.5)
    pyautogui.click(93, 52)
    time.sleep(1.0)
    
    # 2. Clic Aceptar (Popup de confirmación)
    # Usamos 975, 515 que es similar al de los otros popups
    print("    - Paso 2: Clic Aceptar (975, 515)")
    pyautogui.moveTo(975, 515, duration=0.5)
    pyautogui.click(975, 515)
    time.sleep(1.0)
    
    # 3. Clic Nuevo
    print("    - Paso 3: Clic Nuevo (27, 61)")
    pyautogui.moveTo(27, 61, duration=0.5)
    pyautogui.click(27, 61)
    time.sleep(2.0)
    
    BotState.is_saving = False

def process_row(index, row):
    razon_social = row.get('Razon Social', 'N/A')
    identificacion = row.get('Identificación', 'N/A')
    
    print(f"\n[*] [Fila {index+1}] {razon_social} | ID: {identificacion}")
    informacion_basica(row)
    facturacion(row)
    contacto(row)
    direccion(row)
    telefono(row)
    correo(row)
    guardar_y_continuar()
    time.sleep(0.5)

def main():
    print("===========================================")
    print("   NUEVO BOT - REVERTIDO A LÓGICA ORIGINAL")
    print("===========================================")
    
    df = read_data_from_excel(EXCEL_FILE)
    if df is None: return

    # Iniciar cronómetro
    start_time = time.time()

    print("\n[*] Tienes 5 segundos para prepararte...")
    time.sleep(5)
    
    active_window = gw.getActiveWindow()
    if active_window:
        BotState.main_window_title = active_window.title
        print(f"[*] Ventana principal: '{BotState.main_window_title}'")
        
    threading.Thread(target=monitor_de_popups, daemon=True).start()
    
    exitosas = 0
    fallidas = 0
    total = len(df)
    
    for index, row in df.iterrows():
        BotState.current_row_index = index + 1
        
        while BotState.pause_requested:
            winsound.Beep(800, 300)
            respuesta = input(f"\n--> ¿Deseas reanudar desde la fila {BotState.current_row_index}? (s/n): ")
            if respuesta.lower() == 's':
                print("Reanudando en 3 segundos...")
                time.sleep(3)
                BotState.pause_requested = False
            else:
                sys.exit(0)

        try:
            process_row(index, row)
            exitosas += 1
            marcar_completado(index + 2, 'OK') 
        except Exception as e:
            fallidas += 1
            if str(e) == "DIGITO_VERIFICACION" or str(e) == "CLIENTE_YA_CREADO":
                marcar_completado(index + 2, 'POPUP')
                es_creado = (str(e) == "CLIENTE_YA_CREADO")
                print(f"    - [!] Popup detectado. Ejecutando escape automático...")
                
                BotState.is_saving = True 
                time.sleep(0.5)
                if es_creado:
                    pyautogui.click(x=853, y=512) # Aceptar aviso
                    time.sleep(0.5)
                    pyautogui.click(x=17, y=60)   # Nuevo
                    time.sleep(0.5)
                    pyautogui.click(x=442, y=228) # Confirmar descarte
                else:
                    pyautogui.click(x=975, y=515) # Cerrar aviso digito
                    time.sleep(0.5)
                    pyautogui.click(x=27, y=61)   # Nuevo
                    time.sleep(0.5)
                    pyautogui.click(x=885, y=520) # Confirmar descarte
                    
                time.sleep(1.5)
                BotState.digito_requested = False
                BotState.creado_requested = False
                BotState.is_saving = False 
                print("    - [!] Saltando fila.\n")
            else:
                marcar_completado(index + 2, 'ERROR')
                print(f"[-] Error en fila {index+1}: {e}")

    print("\n===========================================")
    print("   EJECUCIÓN FINALIZADA")
    
    # Calcular duración
    end_time = time.time()
    duracion_total = end_time - start_time
    minutos = int(duracion_total // 60)
    segundos = int(duracion_total % 60)
    
    print(f"   Resultados: {exitosas} OK, {fallidas} Fallidos de {total}")
    print(f"   Tiempo total: {minutos}m {segundos}s")
    if total > 0:
        promedio = duracion_total / total
        print(f"   Promedio por registro: {promedio:.2f}s")
    print("===========================================")
    
    # Sonido suave y largo al finalizar (Arpegio ascendente suave)
    print("[*] Proceso completado. Notificando con sonido...")
    try:
        frecuencias = [440, 554, 659, 880] # La mayor: A4, C#5, E5, A5
        for f in frecuencias:
            winsound.Beep(f, 400)
            time.sleep(0.05)
        winsound.Beep(1108, 1000) # Nota final larga (C#6)
    except:
        pass

if __name__ == "__main__":
    main()
