import pyautogui
import pandas as pd
import openpyxl
import time
import sys
import winsound
import pygetwindow as gw
import threading

# ==========================================
# CONFIGURACIÓN INICIAL
# ==========================================
# Ruta de tu archivo Excel (debe tener las columnas correctas)
EXCEL_FILE = 'datos.xlsx'

# Nombre de la hoja de cálculo (dejar en 0 si es la primera hoja)
SHEET_NAME = 0 

# Pausa estándar entre cada acción del bot (en segundos)
pyautogui.PAUSE = 0.2  # Reducido de 0.3 → ~2s menos por registro (20 acciones × 0.1s)

# Medida de seguridad: Si mueves el mouse a una de las esquinas de la pantalla, el bot se detendrá
pyautogui.FAILSAFE = True

# ==========================================
# VARIABLES GLOBALES
# ==========================================
class BotState:
    pause_requested = False
    digito_requested = False
    creado_requested = False
    current_row_index = 0
    main_window_title = ""
    is_saving = False
    modo_alternativo = False

# ==========================================
# KEY LOGGER (registro de teclas con timestamp)
# ==========================================
_key_log = None  # Se abre en main()

def _log_key(accion, detalle):
    """Escribe una entrada en el log de teclas con timestamp de milisegundos."""
    if _key_log:
        ts = time.strftime('%H:%M:%S') + f'.{int(time.time() * 1000) % 1000:03d}'
        _key_log.write(f'[{ts}] {accion:<8} {detalle}\n')
        _key_log.flush()

# --- Referencias originales de pyautogui ---
_orig_press  = pyautogui.press
_orig_write  = pyautogui.write
_orig_click  = pyautogui.click
_orig_hotkey = pyautogui.hotkey
_orig_moveTo = pyautogui.moveTo

def _p_press(key, *a, **kw):
    _log_key('PRESS', repr(key))
    return _orig_press(key, *a, **kw)

def _p_write(text, *a, **kw):
    _log_key('WRITE', repr(str(text)))
    return _orig_write(text, *a, **kw)

def _p_click(*a, **kw):
    x = kw.get('x', a[0] if a else '?')
    y = kw.get('y', a[1] if len(a) > 1 else '?')
    _log_key('CLICK', f'({x}, {y})')
    return _orig_click(*a, **kw)

def _p_hotkey(*a, **kw):
    _log_key('HOTKEY', '+'.join(str(k) for k in a))
    return _orig_hotkey(*a, **kw)

def _p_moveTo(*a, **kw):
    x = a[0] if a else kw.get('x', '?')
    y = a[1] if len(a) > 1 else kw.get('y', '?')
    _log_key('MOVE', f'({x}, {y})')
    return _orig_moveTo(*a, **kw)

def activar_key_logger():
    """Reemplaza las funciones de pyautogui con versiones que loguean cada acción."""
    pyautogui.press  = _p_press
    pyautogui.write  = _p_write
    pyautogui.click  = _p_click
    pyautogui.hotkey = _p_hotkey
    pyautogui.moveTo = _p_moveTo

# ==========================================
# FUNCIONES DEL BOT
# ==========================================

def monitor_de_popups():
    """
    Hilo en segundo plano que vigila constantemente la ventana activa.
    Si el título cambia a uno inesperado, activa la alerta de pausa.
    """
    while True:
        try:
            active_window = gw.getActiveWindow()
            if active_window and BotState.main_window_title and not BotState.is_saving:
                current_title = active_window.title
                
                # Ignorar si no tiene titulo, o si es la ventana principal, o la consola de Python, o si es la misma ventana de Abako cambiando su título
                if current_title and current_title != BotState.main_window_title and "abako" not in current_title.lower() and "cmd.exe" not in current_title.lower() and "python" not in current_title.lower():
                    if "digito" in current_title.lower() or "dígito" in current_title.lower():
                        if not getattr(BotState, 'digito_requested', False):
                            print(f"\n    [!] Detectado popup: '{current_title}'. Solicitando escape automático...")
                            BotState.digito_requested = True
                    elif "planilla nuevo cliente" in current_title.lower() or "recuperar su" in current_title.lower():
                        if not getattr(BotState, 'creado_requested', False):
                            print(f"\n    [!] Detectado popup de cliente existente: '{current_title}'. Solicitando escape automático...")
                            BotState.creado_requested = True
                    else:
                        # Comprobamos si no está ya pausado para no hacer spam en consola
                        if not BotState.pause_requested:
                            print("\n" + "!"*60)
                            print(f"[ALERTA] Se detectó una ventana emergente inesperada: '{current_title}'")
                            print(f"[ALERTA] Leyendo actualmente la Fila de Excel {BotState.current_row_index}.")
                            print("[ALERTA] Escribe 's' y presiona ENTER para continuar a la siguiente fila.")
                            print("!"*60 + "\n")
                            BotState.pause_requested = True
                            # 🔔 Tres beeps de alerta para avisar al usuario
                            for _ in range(3):
                                winsound.Beep(1000, 400)  # 1000 Hz, 400ms
                                time.sleep(0.15)
                        
        except Exception:
            pass # Ignorar errores si la ventana se cierra muy rápido al intentar leerla
            
        import time
        time.sleep(0.1)

def check_pause():
    """
    Verifica si se ha solicitado una pausa. Si es así, lanza una excepción
    para interrumpir el flujo del bot inmediatamente.
    """
    if getattr(BotState, 'digito_requested', False):
        raise Exception("DIGITO_VERIFICACION")
        
    if getattr(BotState, 'creado_requested', False):
        raise Exception("CLIENTE_YA_CREADO")
        
    if BotState.pause_requested:
        raise Exception("Interrupción de seguridad: Popup inesperado detectado.")

def smart_sleep(duration):
    """
    Duerme por un tiempo especificado, pero verificando constantemente si hay
    popups para interrumpir el proceso de ser necesario.
    """
    import time
    slices = int(duration / 0.1)
    remainder = duration % 0.1
    
    for _ in range(slices):
        check_pause()
        time.sleep(0.1)
        
    if remainder > 0:
        check_pause()
        time.sleep(remainder)

def read_data_from_excel(file_path):
    """
    Lee los datos desde un archivo Excel y retorna un DataFrame de Pandas.
    """
    try:
        print(f"[*] Leyendo archivo Excel: {file_path}...")
        df = pd.read_excel(file_path, sheet_name=SHEET_NAME)
        print(f"[+] Se cargaron {len(df)} filas correctamente.")
        return df
    except Exception as e:
        print(f"[-] Error al leer el archivo Excel: {e}")
        print("Asegúrate de que 'datos.xlsx' existe, tiene los encabezados que indicaste y no está abierto en otro programa.")
        sys.exit(1)

def marcar_completado(excel_row_num, valor='X'):
    """
    Abre el archivo Excel con openpyxl y escribe el 'valor' en la columna 'Completado'
    de la fila indicada (excel_row_num es el número de fila real en Excel, incluyendo encabezado).
    Si la columna 'Completado' no existe, la crea automáticamente al final.
    """
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        # Obtener la hoja correcta
        if isinstance(SHEET_NAME, int):
            ws = wb.worksheets[SHEET_NAME]
        else:
            ws = wb[SHEET_NAME]

        # Buscar si ya existe la columna 'Completado' en la fila de encabezados (fila 1)
        col_completado = None
        for cell in ws[1]:
            if cell.value and str(cell.value).strip().lower() == 'completado':
                col_completado = cell.column
                break

        # Si no existe, crearla en la siguiente columna disponible
        if col_completado is None:
            col_completado = ws.max_column + 1
            ws.cell(row=1, column=col_completado, value='Completado')
            print(f"    [Excel] Se creó la columna 'Completado' en la columna {col_completado}.")

        # Escribir la marca en la fila correspondiente
        ws.cell(row=excel_row_num, column=col_completado, value=valor)
        wb.save(EXCEL_FILE)
        print(f"    [Excel] ✅ Fila {excel_row_num} marcada como completada en el Excel.")
    except Exception as e:
        print(f"    [Excel] ⚠️ No se pudo marcar la fila {excel_row_num} en el Excel: {e}")


def limpiar_texto(texto):
    """
    Elimina saltos de línea, tabulaciones y cualquier caracter de control
    de un string para que pyautogui no los interprete como teclas especiales.
    Sólo deja texto imprimible plano.
    """
    import unicodedata
    if not texto:
        return ""
    # Reemplazar saltos de línea y tabs por espacio
    texto = texto.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
    # Eliminar caracteres de control (categoría Unicode 'C' excepto espacio normal)
    texto = ''.join(c for c in texto if not unicodedata.category(c).startswith('C'))
    # Colapsar múltiples espacios en uno solo y quitar espacios al inicio/final
    texto = ' '.join(texto.split())
    return texto

def informacion_basica(row):
    """
    Llena la sección de Información Básica del formulario (Pasos 1 al 7).
    """
    print("\n--- [ INFORMACIÓN BÁSICA ] ---")
    
    # EXTRACCIÓN Y LIMPIEZA DE DATOS DE EXCEL
    # limpiar_texto() elimina \n, \r, \t y otros caracteres especiales que rompen el flujo
    dato_razon_social   = limpiar_texto(str(row.get('Razon Social', '')))
    dato_nombre_comun   = limpiar_texto(str(row.get('Nombre Común', '')))
    dato_identificacion = limpiar_texto(str(row.get('Identificación', '')))
    if dato_identificacion.endswith('.0'): dato_identificacion = dato_identificacion[:-2]
    dato_canal_texto    = limpiar_texto(str(row.get('Canal', '')))
    dato_segmento_texto = limpiar_texto(str(row.get('Segmento', '')))
    dato_asesor_texto   = limpiar_texto(str(row.get('Asesor', '')))

    # Si hay valores nulos o vacíos en pandas (nan), los dejamos como string vacío
    if dato_razon_social.lower()   == 'nan': dato_razon_social   = ""
    if dato_nombre_comun.lower()   == 'nan': dato_nombre_comun   = ""
    if dato_identificacion.lower() == 'nan': dato_identificacion = ""
    if dato_canal_texto.lower()    == 'nan': dato_canal_texto    = ""
    if dato_segmento_texto.lower() == 'nan': dato_segmento_texto = ""
    if dato_asesor_texto.lower()   == 'nan': dato_asesor_texto   = ""    
    dato_forma_pago = str(row.get('Forma de pago', ''))
    dato_cupo = str(row.get('Cupo', ''))
    dato_cuotas = str(row.get('Cuotas', ''))
    dato_vencimiento = str(row.get('Vencimiento', ''))
    dato_lista_precios = str(row.get('Lista de Precios', ''))
    
    if dato_forma_pago.lower() == 'nan': dato_forma_pago = ""
    if dato_cupo.lower() == 'nan': dato_cupo = ""
    if dato_cuotas.lower() == 'nan': dato_cuotas = ""
    if dato_vencimiento.lower() == 'nan': dato_vencimiento = ""
    if dato_lista_precios.lower() == 'nan': dato_lista_precios = ""

    # Lógica de teclas para el selector de Canal
    dato_canal = ""
    canal_upper = dato_canal_texto.upper()
    if "SUPERMERCADO" in canal_upper:
        dato_canal = "sss"
    elif "MINORISTA" in canal_upper or "MINIMERCADO" in canal_upper:
        dato_canal = "mm"
    elif "MAYORISTA" in canal_upper:
        dato_canal = "m"
    elif "TIENDA" in canal_upper:
        dato_canal = "t"
    elif canal_upper:
        dato_canal = canal_upper[0] # Por defecto, la primera letra
    
    # Lógica de teclas para el selector de Segmento
    dato_segmento = ""
    segmento_upper = dato_segmento_texto.upper()
    if "SUPERMERCADO" in segmento_upper:
        dato_segmento = "ssss"
    elif "MINORISTA" in segmento_upper or "MINIMERCADO" in segmento_upper:
        dato_segmento = "mm"
    elif "MAYORISTA" in segmento_upper:
        dato_segmento = "m"
    elif "TIENDA" in segmento_upper:
        dato_segmento = "t"
    elif segmento_upper:
        dato_segmento = segmento_upper[0] # Por defecto, la primera letra

    # Paso 1: Clic inicial en Razón Social y escribir
    print("    - Paso 1: Escribiendo Razón Social")
    time.sleep(0.5)  # Reducido de 1.0s
    pyautogui.moveTo(515, 115, duration=0.5)
    pyautogui.click(x=515, y=115)
    pyautogui.click(x=515, y=115)
    time.sleep(0.5)
    pyautogui.write(dato_razon_social)
    time.sleep(1.0)  # Reducido de 1.5s

    # Paso 2: ENTER+ENTER → Nombre Común y escribir
    # El log manual del usuario confirma: siempre ENTER+ENTER para navegar, nunca TAB
    print("    - Paso 2: Presionando ENTER y escribiendo Nombre Común")
    pyautogui.press("enter")
    time.sleep(0.2)
    pyautogui.press("enter")
    time.sleep(0.8)  # Espera a que el campo Nombre Común esté activo
    pyautogui.write(dato_nombre_comun)
    time.sleep(2.0)  # Espera a que ABAKO procese el texto

    # Paso 3: ENTER+ENTER → Tipo de Documento, luego 'c' + ENTER para confirmar Cédula
    # El log manual del usuario muestra: ~155ms entre los dos ENTERs, ~872ms antes de 'c'
    print("    - Paso 3: Presionando ENTER y Seleccionando Tipo de Documento (C + Enter)")
    pyautogui.press("enter")
    if not BotState.modo_alternativo:
        time.sleep(0.2)
        pyautogui.press("enter")
    time.sleep(1.0)  # ~872ms en el log manual → usamos 1.0s para dar un poco más de margen
    pyautogui.press("c")  # Minúscula directa para filtrar "Cédula"
    time.sleep(1.5)  # Espera a que ABAKO filtre y seleccione la opción Cédula
    pyautogui.press("enter")  # <-- Confirma Cédula
    time.sleep(1.0)  # Espera a que pase al campo Identificación

    # Paso 4: Escribir Identificación (el campo ya está activo tras el enter anterior)
    print("    - Paso 4: Escribiendo Identificación")
    time.sleep(1.0)  # Seguridad extra: espera a que el foco esté en Identificación
    pyautogui.write(dato_identificacion)
    time.sleep(0.5)
    pyautogui.press("tab")
    
    # Pausa vigilada post-identificación: si aparece un popup (ej: duplicado), pausar inmediatamente
    print("    - Paso 4.1: Esperando y verificando popups post-identificación...")
    smart_sleep(1.0)  # Reducido de 1.5s
    check_pause()
    
    # Paso 5: TAB hacia Asesor y seleccionar opción
    print(f"    - Paso 5: Presionando TAB y Seleccionando Asesor ({dato_asesor_texto})")
    pyautogui.press("tab")
    time.sleep(0.5)  # Reducido de 1.0s
    
    dato_asesor_corto = dato_asesor_texto[:3] if dato_asesor_texto else ""
    if dato_asesor_corto:
        pyautogui.write(dato_asesor_corto)
    time.sleep(0.3)  # Reducido de 0.5s
    
    # Confirmar Asesor con Enter
    pyautogui.press("enter")
    time.sleep(0.3)  # Reducido de 0.5s

    # Paso 6: TAB hacia Canal y seleccionar opción
    print(f"    - Paso 6: Presionando TABs y Seleccionando Canal ({dato_canal_texto} -> {dato_canal} + Enter)")
    time.sleep(0.5)  # Reducido de 1.0s
    
    if dato_canal:
        pyautogui.write(dato_canal, interval=0.15)
    time.sleep(0.3)  # Reducido de 0.5s
    
    # Después de las teclas especiales, indicaste que viene un enter
    pyautogui.press("enter")
    time.sleep(0.3)  # Reducido de 0.5s
    
    # Paso 7: TAB hacia Segmento y seleccionar opción
    print(f"    - Paso 7: Presionando TAB y Seleccionando Segmento ({dato_segmento_texto})")
    time.sleep(0.5)  # Reducido de 1.0s
    
    if dato_segmento:
        pyautogui.write(dato_segmento, interval=0.15)
    time.sleep(0.3)  # Reducido de 0.5s
    
    # Después de las teclas especiales viene un enter
    pyautogui.press("enter")
    time.sleep(0.3)  # Reducido de 0.5s


def facturacion(row):
    """
    Llena la sección de Facturación.
    """
    print("\n--- [ FACTURACIÓN ] ---")
    
    dato_forma_pago = str(row.get('Forma de pago', ''))
    dato_cupo = str(row.get('Cupo', ''))
    dato_cuotas = str(row.get('Cuotas', ''))
    dato_vencimiento = str(row.get('Vencimiento', ''))
    dato_lista_precios = str(row.get('Lista de Precios', ''))
    
    # Manejar NaN
    if dato_forma_pago.lower() == 'nan': dato_forma_pago = ""
    if dato_cupo.lower() == 'nan': dato_cupo = ""
    if dato_cuotas.lower() == 'nan': dato_cuotas = ""
    if dato_vencimiento.lower() == 'nan': dato_vencimiento = ""
    if dato_lista_precios.lower() == 'nan': dato_lista_precios = ""
    
    # Lógica Lista de Precios
    dato_lista_corto = dato_lista_precios[:3] if dato_lista_precios else ""

    # Paso 1: Clic Normal en Tipo de Tercero
    print("    - Paso 1: Clic Inicial (389, 414)")
    pyautogui.moveTo(389, 414, duration=0.5)
    pyautogui.click(x=389, y=414)
    smart_sleep(0.1)
    
    # Si NO es de contado, llenamos los créditos
    if "CONTADO" not in dato_forma_pago.upper():
        print("    - Paso 2: Clic en Forma de Pago Crédito (674, 458)")
        pyautogui.moveTo(674, 458, duration=0.5)
        pyautogui.click(x=674, y=458)
        smart_sleep(0.1)
        
        # Presionar C y Enter
        pyautogui.write("c")
        smart_sleep(0.1)
        pyautogui.press("enter")
        smart_sleep(0.1)
        
        # Escribir cupo, doble enter
        print(f"    - Paso 3: Escribiendo Cupo ({dato_cupo}) y Presionando Enter x2")
        pyautogui.write(dato_cupo)
        smart_sleep(0.1)
        pyautogui.press("enter")
        smart_sleep(0.1)
        pyautogui.press("enter")
        smart_sleep(0.1)
        
        # Escribir numero de cuotas, doble enter
        print(f"    - Paso 4: Escribiendo Cuotas ({dato_cuotas}) y Presionando Enter x2")
        pyautogui.write(dato_cuotas)
        smart_sleep(0.1)
        pyautogui.press("enter")
        smart_sleep(0.1)
        
        # si es 30 (Días), presionar M y enter
        vencimiento_num = ''.join(filter(str.isdigit, dato_vencimiento))
        if vencimiento_num == "30":
            print("    - Paso 4.1: Vencimiento 30 días detectado. Presionando M y Enter.")
            pyautogui.write("m")
            smart_sleep(0.1)
            pyautogui.press("enter")
            smart_sleep(0.1)
    else:
        print("    --> Forma de pago Contado detectada. Omitiendo pasos de crédito.")

    # (Llegamos al punto 🫡)
    # Paso 5: Clic Normal Lista de Precios
    check_pause()
    print(f"    - Paso 5: Clic en Lista de Precios ({dato_lista_precios}) -> ({dato_lista_corto} + Enter)")
    pyautogui.moveTo(1054, 461, duration=0.5)
    pyautogui.click(x=1054, y=461)
    smart_sleep(0.1)
    
    if dato_lista_corto:
        pyautogui.write(dato_lista_corto, interval=0.15)
    smart_sleep(0.1)
    pyautogui.press("enter")
    smart_sleep(0.1)
    
def contacto(row):
    """
    Llena la sección de Contacto.
    """
    print("\n--- [ CONTACTO ] ---")
    
    dato_nombre_contacto = str(row.get('Nombre Contacto', ''))
    dato_apellido_contacto = str(row.get('Apellido Contacto', ''))
    dato_identificacion_contacto = str(row.get('Identificación Contacto', ''))
    if dato_identificacion_contacto.endswith('.0'): dato_identificacion_contacto = dato_identificacion_contacto[:-2]
    
    if dato_nombre_contacto.lower() == 'nan': dato_nombre_contacto = ""
    if dato_apellido_contacto.lower() == 'nan': dato_apellido_contacto = ""
    if dato_identificacion_contacto.lower() == 'nan': dato_identificacion_contacto = ""

    # Paso 1: Clic Normal en Contacto
    check_pause()
    print("    - Paso 1: Clic Inicial (491, 416)")
    pyautogui.moveTo(491, 416, duration=0.5)
    pyautogui.click(x=491, y=416)
    smart_sleep(0.1)
    
    # Escribe nombre contacto y Enter
    print(f"    - Paso 2: Borrando y Escribiendo Nombre Contacto ({dato_nombre_contacto}) y Presionando Enter")
    for _ in range(4):
        pyautogui.hotkey('ctrl', 'delete')
    pyautogui.write(dato_nombre_contacto)
    smart_sleep(0.1)
    pyautogui.press("enter")
    smart_sleep(0.1)
    
    # Escribe apellidos contacto y enter
    print(f"    - Paso 3: Borrando y Escribiendo Apellido Contacto ({dato_apellido_contacto}) y Presionando Enter")
    for _ in range(4):
        pyautogui.hotkey('ctrl', 'delete')
    pyautogui.write(dato_apellido_contacto)
    smart_sleep(0.1)
    pyautogui.press("enter")
    smart_sleep(0.1)
    
    # Escribe indentificacion contacto (Sin enter al final, según lo especificado)
    print(f"    - Paso 4: Escribiendo Identificación Contacto ({dato_identificacion_contacto})")
    pyautogui.write(dato_identificacion_contacto)
    smart_sleep(0.1)

def direccion(row):
    """
    Llena la sección de Dirección.
    """
    print("\n--- [ DIRECCIÓN ] ---")
    
    dato_departamento = str(row.get('Departamento', ''))
    dato_ciudad = str(row.get('Ciudad', ''))
    dato_direccion = str(row.get('Direccion', ''))
    
    if dato_departamento.lower() == 'nan': dato_departamento = ""
    if dato_ciudad.lower() == 'nan': dato_ciudad = ""
    if dato_direccion.lower() == 'nan': dato_direccion = ""

    # Lógica de Departamento (2 letras)
    dato_departamento_corto = dato_departamento[:2] if dato_departamento else ""
    
    # Lógica de Ciudad (3 letras)
    dato_ciudad_corto = dato_ciudad[:3] if dato_ciudad else ""

    print("    - Paso 1: Clic (556, 408)")
    check_pause()
    pyautogui.moveTo(556, 408, duration=0.5)
    pyautogui.click(x=556, y=408)
    smart_sleep(0.1)
    
    print("    - Paso 2: Clic en Departamento (398, 501)")
    pyautogui.moveTo(398, 501, duration=0.5)
    pyautogui.click(x=398, y=501)
    smart_sleep(0.1)
    
    print(f"    - Paso 2.1: Escribiendo Departamento ({dato_departamento_corto}) + Enter")
    if dato_departamento_corto:
        pyautogui.write(dato_departamento_corto, interval=0.15)
    smart_sleep(0.1)
    pyautogui.press("enter")
    smart_sleep(0.1)
    
    print(f"    - Paso 3: Escribiendo Ciudad ({dato_ciudad_corto}) + Enter")
    if dato_ciudad_corto:
        pyautogui.write(dato_ciudad_corto, interval=0.15)
    smart_sleep(0.1)
    pyautogui.press("enter")
    smart_sleep(0.1)

    print("    - Paso 3.1: Escribiendo Barrio (sin) + Enter")
    pyautogui.write("sin", interval=0.15)
    smart_sleep(0.1)
    pyautogui.press("enter")
    smart_sleep(0.1)
    
    print("    - Paso 4: Clic en tipo de vía (745, 477)")
    pyautogui.moveTo(745, 477, duration=0.5)
    pyautogui.click(x=745, y=477)
    smart_sleep(0.1)
    
    print("    - Paso 4.1: Escribiendo 'ave' + Enter x2")
    pyautogui.write("ave", interval=0.15)
    smart_sleep(0.1)
    pyautogui.press("enter")
    smart_sleep(0.1)

    
    print(f"    - Paso 5: Escribiendo Dirección completa ({dato_direccion})")
    pyautogui.write(dato_direccion)
    smart_sleep(0.1)
    pyautogui.press("enter")
    smart_sleep(0.1)

    pyautogui.write(".", interval=0.15)

def telefono(row):
    """
    Llena la sección de Teléfono.
    """
    print("\n--- [ TELÉFONO ] ---")
    
    dato_telefono = str(row.get('Telefono', ''))
    if dato_telefono.endswith('.0'): dato_telefono = dato_telefono[:-2]
    if dato_telefono.lower() == 'nan': dato_telefono = ""

    print("    - Paso 1: Clic (693, 410)")
    check_pause()
    pyautogui.moveTo(693, 410, duration=0.5)
    pyautogui.click(x=693, y=410)
    smart_sleep(0.1)
    
    print("    - Paso 2: Clic (45, 54)")
    pyautogui.moveTo(45, 54, duration=0.5)
    pyautogui.click(x=45, y=54)
    smart_sleep(0.1)
    
    print("    - Paso 3: Clic (676, 475)")
    pyautogui.moveTo(676, 475, duration=0.5)
    pyautogui.click(x=676, y=475)
    smart_sleep(0.1)
    
    print(f"    - Paso 4: Clic y Escribir Teléfono ({dato_telefono}) en (697, 467)")
    pyautogui.moveTo(697, 467, duration=0.5)
    pyautogui.click(x=697, y=467)
    smart_sleep(0.1)
    pyautogui.write(dato_telefono)
    smart_sleep(0.1)

def guardar_y_continuar():
    """
    Guarda el registro actual y prepara el formulario para la siguiente fila.
    """
    print("\n--- [ GUARDAR Y CONTINUAR ] ---")
    BotState.is_saving = True
    
    print("    - Paso 1: Clic Guardar (93, 52)")
    pyautogui.moveTo(93, 52, duration=0.5)
    pyautogui.click(x=93, y=52)
    time.sleep(0.3)  # Reducido de 0.5s
    
    print("    - Paso 2: Clic Aceptar (948, 517)")
    pyautogui.moveTo(948, 517, duration=0.5)
    pyautogui.click(x=948, y=517)
    time.sleep(0.3)  # Reducido de 0.5s
    
    print("    - Paso 3: Clic Nuevo (26, 55)")
    pyautogui.moveTo(26, 55, duration=0.5)
    pyautogui.click(x=26, y=55)
    time.sleep(0.3)  # Reducido de 0.5s
    
    # Pausa para que ABAKO cargue el formulario en blanco
    time.sleep(1.5)  # Reducido de 2.0s
    
    active_w = gw.getActiveWindow()
    if active_w and "cmd.exe" not in active_w.title.lower() and "python" not in active_w.title.lower():
         BotState.main_window_title = active_w.title
         
    BotState.is_saving = False
    
    print("    - Pausa extra antes de continuar con el siguiente...")
    time.sleep(0.5)  # Reducido de 1.0s
    
def process_row(row):
    """
    Contiene la lógica de navegación y clics para UNA fila de datos del Excel.
    """
    print(f"[*] Procesando fila: {row.to_dict()}")
    
    # 1. Llamar al flujo de Información Básica
    informacion_basica(row)
    
    # 2. Llamar al flujo de Facturación
    facturacion(row)
    
    # 3. Llamar al flujo de Contacto
    contacto(row)
    
    # 4. Llamar al flujo de Dirección
    direccion(row)
    
    # 5. Llamar al flujo de Teléfono
    telefono(row)

    # 6. Guardar y preparar la siguiente fila
    guardar_y_continuar()

    print("[+] Fila procesada con éxito.\n")


def main():
    global _key_log
    
    if "--alter" in sys.argv:
        BotState.modo_alternativo = True
        
    print("===========================================")
    print("    INICIANDO BOT DE AUTOMATIZACIÓN")
    if BotState.modo_alternativo:
        print("    [MODO ALTERNATIVO ACTIVO]")
    print("===========================================")
    print("Asegúrate de preparar la pantalla en el formulario de registro.")
    print("Tienes 5 segundos para prepararte y cambiar a la ventana deseada...")
    time.sleep(5)

    # Métricas de tiempo de inicio
    start_time = time.time()

    # Abrir log de teclas y activar el interceptor
    _key_log = open('log_teclas.txt', 'a', encoding='utf-8')
    _key_log.write('\n' + '='*60 + '\n')
    _key_log.write(f'  KEY LOG: {time.strftime("%Y-%m-%d %H:%M:%S")}\n')
    _key_log.write('='*60 + '\n')
    _key_log.flush()
    activar_key_logger()
    print("[*] Key-logger activado → log_teclas.txt")

    # 1. Leer datos del Excel
    df = read_data_from_excel(EXCEL_FILE)
    total_filas = len(df)
    exitosas = 0
    fallidas = 0
    registros_fallidos = []

    # Abrir el archivo de log en modo append (agrega sin borrar ejecuciones anteriores)
    LOG_FILE = 'log_ejecucion.txt'
    log = open(LOG_FILE, 'a', encoding='utf-8')
    log.write("\n" + "="*60 + "\n")
    log.write(f"  NUEVA EJECUCIÓN: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
    log.write("="*60 + "\n")
    log.flush()
    print(f"[*] Log de ejecución guardándose en: {LOG_FILE}")
    
    # Capturar la ventana principal antes de iniciar
    active_window = gw.getActiveWindow()
    if active_window:
        BotState.main_window_title = active_window.title
        print(f"[*] Ventana principal fijada como: '{BotState.main_window_title}'")
        
    # Iniciar el hilo del monitor de popups
    monitor_thread = threading.Thread(target=monitor_de_popups, daemon=True)
    monitor_thread.start()
    
    # 2. Iterar sobre TODAS las filas del Excel (sin límite)
    for index, row in df.iterrows():
        # Actualizar índice global para el hilo del monitor
        BotState.current_row_index = index + 2 
        
        # Bucle de espera si se detectó una pausa
        while BotState.pause_requested:
            # Beep recordatorio cada vez que pide confirmación
            winsound.Beep(800, 300)
            print("\n--> ESPERANDO ORDEN DE REANUDA...")
            respuesta = input(f"--> ¿Deseas reanudar desde la fila {BotState.current_row_index}? (s/n): ")
            if respuesta.lower() == 'n':
                print("[!] Ejecución detenida por el usuario.")
                sys.exit(0)
            elif respuesta.lower() == 's':
                print("\nReanudando en 3 segundos. Por favor, vuelve a la ventana principal...")
                winsound.Beep(600, 200)  # beep suave de confirmación
                time.sleep(3)
                BotState.pause_requested = False
            else:
                print("Entrada no válida. Escribe 's' para sí o 'n' para no.")

        identificacion_actual = str(row.get('Identificación', 'Desconocido'))
        if identificacion_actual.endswith('.0'): identificacion_actual = identificacion_actual[:-2]
        nombre_actual = str(row.get('Nombre Común', 'Desconocido'))
        
        print(f"--- Iniciando iteración {index + 1} de {total_filas} ---")
        print(f"    🔵 ID: {identificacion_actual}  |  Nombre: {nombre_actual}")
        try:
            process_row(row)
            exitosas += 1
            # Marcar la fila como completada en el Excel (índice 0 de pandas + 1 de encabezado + 1 = +2)
            marcar_completado(index + 2)
            # Registrar éxito en el log
            log.write(f"[✅ OK]  Fila {index + 2:>4} | ID: {identificacion_actual:<20} | {nombre_actual}\n")
            log.flush()
        except Exception as e:
            if str(e) == "DIGITO_VERIFICACION" or str(e) == "CLIENTE_YA_CREADO":
                # Resetear banderas
                es_creado = (str(e) == "CLIENTE_YA_CREADO")
                if es_creado:
                    BotState.creado_requested = False
                    print("    - [!] Excepción de Cliente Existente atrapada. Ejecutando escape automático...")
                else:
                    BotState.digito_requested = False
                    print("    - [!] Excepción de Dígito de Verificación atrapada. Ejecutando escape automático...")
                
                # Desactivar temporalmente monitor de popups mientras escapamos
                BotState.is_saving = True 
                
                time.sleep(0.5)
                if es_creado:
                    # Paso 1 [CLICK] -> 853, 512 (No en el popup cliente creado)
                    pyautogui.click(x=853, y=512)
                    time.sleep(0.5)
                    # Paso 2 [CLICK] -> 17, 60 (Botón Nuevo)
                    pyautogui.click(x=17, y=60)
                    time.sleep(0.5)
                    # Paso 3 [CLICK] -> 442, 228 (Confirmar)
                    pyautogui.click(x=442, y=228)
                else:
                    # Paso 1 [CLICK] -> 975, 515 (Cerrar / No en el popup de digito verif)
                    pyautogui.click(x=975, y=515)
                    time.sleep(0.5)
                    # Paso 2 [CLICK] -> 27, 61 (Botón Nuevo)
                    pyautogui.click(x=27, y=61)
                    time.sleep(0.5)
                    # Paso 3 [CLICK] -> 885, 520 (Confirmar descartar cambios)
                    pyautogui.click(x=885, y=520)
                    
                time.sleep(1.5) # Esperar a que recargue el formulario
                
                # Restaurar foco de ABAKO
                active_w = gw.getActiveWindow()
                if active_w and "cmd.exe" not in active_w.title.lower() and "python" not in active_w.title.lower():
                     BotState.main_window_title = active_w.title
                
                BotState.is_saving = False 
                
                if es_creado:
                    marcar_completado(index + 2, valor="CREADO")
                    exitosas += 1 # Contar como procesada (saltada)
                    log.write(f"[⚠️ SKIP] Fila {index + 2:>4} | ID: {identificacion_actual:<20} | SALTADA POR EXISTENTE (CREADO)\n")
                    log.flush()
                    print("[!] Fila saltada exitosamente y marcada como 'CREADO'. Continuando...\n")
                else:
                    marcar_completado(index + 2, valor="!DIGITO!")
                    exitosas += 1 # Contar como procesada (saltada)
                    log.write(f"[⚠️ SKIP] Fila {index + 2:>4} | ID: {identificacion_actual:<20} | SALTADA POR DÍGITO VERIF.\n")
                    log.flush()
                    print("[!] Fila saltada exitosamente y marcada como '!DIGITO!'. Continuando...\n")
                
            else:
                print(f"[-] Error procesando la fila {index + 1}: {e}")
                print("[!] Continuando con la siguiente fila...\n")
                fallidas += 1
            
            # Guardamos el detalle del error para el reporte final
            registros_fallidos.append({
                "Fila Excel": index + 2,
                "Identificación": identificacion_actual,
                "Nombre": nombre_actual,
                "Error": str(e)
            })
            # Registrar fallo en el log EN TIEMPO REAL
            log.write(f"[❌ FAIL] Fila {index + 2:>4} | ID: {identificacion_actual:<20} | {nombre_actual} | ERROR: {e}\n")
            log.flush()
            continue
            
    end_time = time.time()
    tiempo_total = end_time - start_time
    promedio = tiempo_total / total_filas if total_filas > 0 else 0

    log.write("\n--- RESUMEN FINAL ---\n")
    log.write(f"Total: {total_filas} | Exitosos: {exitosas} | Fallidos: {fallidas}\n")
    log.write(f"Tiempo total: {tiempo_total:.2f}s | Promedio: {promedio:.2f}s por registro\n")
    log.close()

    print("===========================================")
    print("    BOT FINALIZADO CON ÉXITO")
    print("===========================================")
    print("    📊 RESUMEN DE EJECUCIÓN:")
    print(f"    - Total registros: {total_filas}")
    print(f"    - Exitosos: ✅ {exitosas}")
    print(f"    - Fallidos: ❌ {fallidas}")
    
    # Reportamos y guardamos errores si los hubo
    if fallidas > 0:
        print("\n    ⚠️ DETALLE DE FALLOS:")
        with open('reporte_errores.txt', 'w', encoding='utf-8') as f:
            f.write("REPORTE DE ERRORES AL PROCESAR EXCEL\n")
            f.write("====================================\n\n")
            for fallo in registros_fallidos:
                detalle = f"Fila Excel: {fallo['Fila Excel']} | ID: {fallo['Identificación']} | Nombre: {fallo['Nombre']} | Error: {fallo['Error']}"
                print(f"      - {detalle}")
                f.write(detalle + "\n")
        print("\n    [!] Se ha guardado un reporte detallado en 'reporte_errores.txt'")
        
    print(f"\n    - Tiempo total: ⏱️ {tiempo_total:.2f} segundos")
    print(f"    - Tiempo promedio por registro: ⏱️ {promedio:.2f} segundos")
    print(f"    - Log completo guardado en: log_ejecucion.txt")
    print(f"    - Log de teclas guardado en: log_teclas.txt")
    print("===========================================")
    _key_log.close()

if __name__ == "__main__":
    main()
